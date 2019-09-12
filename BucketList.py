from enum import Enum
import boto3
from openpyxl import Workbook
from Util.S3Repository import Repository, FileObject, FileVersion

class OutputType(Enum):
    StandardOutput = 1
    TextFile = 2
    Excel = 3


class BucketOutput:
    def __init__(self, output_type, output_header, show_versions, filename=None):
        self._type = output_type
        self._filename = filename
        self._output_header = output_header
        self._show_versions = show_versions

    @property
    def type(self):
        return self._type

    @property
    def filename(self):
        return self._filename

    @property
    def output_header(self):
        return self._output_header

    @property
    def show_versions(self):
        return self._show_versions


def output_file_objects(bucket, output):
    if output.show_versions:
        objects = bucket.object_versions.all()
        header = 'key,id,is_latest,size,last_modified,storage_class,version_id'
    else:
        objects = bucket.objects.all()
        header = 'key,size,last_modified,storage_class'
    if output.type == OutputType.StandardOutput:
        if output.output_header:
            print(header)
        for o in objects:
            print(get_file_object_output(o, output))

    if output.type == OutputType.TextFile:
        with open(output.filename, mode='w') as f:
            if output.output_header:
                f.write(header + '\n')
            for o in objects:
                f.write(get_file_object_output(o, output) + '\n')

    if output.type == OutputType.Excel:
        wb = Workbook()
        ws = wb.active
        if output.output_header:
            header_items = header.split(sep=',')
            for n in range(len(header_items)):
                ws.cell(column=n + 1, row=1, value=header_items[n])
        row = 2
        if output.show_versions:
            c = boto3.client('s3')
            for o in objects:
                ws.cell(row=row, column=1, value=o.key)
                ws.cell(row=row, column=2, value=o.id)
                ws.cell(row=row, column=3, value=o.is_latest)
                ws.cell(row=row, column=4, value=o.size)
                ws.cell(row=row, column=5, value=o.last_modified)
                ws.cell(row=row, column=6, value=o.storage_class)
                ws.cell(row=row, column=7, value=o.version_id)
                try:
                    response = c.head_object(Bucket='jea-scratch', Key=o.key)
                    if 'DeleteMarker' in response:
                        delete_marker = response['DeleteMarker']
                        print(delete_marker)
                except Exception as e:
                    print(e)
                row += 1
        else:
            for o in objects:
                ws.cell(row=row, column=1, value=o.key)
                ws.cell(row=row, column=2, value=o.size)
                ws.cell(row=row, column=3, value=o.last_modified)
                ws.cell(row=row, column=4, value=o.storage_class)
                row += 1
        wb.save(output.filename)


def get_file_object_output(o, output):
    if output.show_versions:
        return f'{o.key},{o.id},{o.is_latest},{o.size},{o.last_modified},{o.storage_class},{o.version_id}'
    else:
        return f'{o.key},{o.id},{o.is_latest},{o.size},{o.last_modified},{o.storage_class},{o.version_id}'


if __name__ == '__main__':

    import argparse

    parser = argparse.ArgumentParser(description='S3 Bucket List')
    parser.add_argument('bucket', help='Bucket name', metavar='bucket-name')
    output_group = parser.add_mutually_exclusive_group()
    output_group.add_argument('-o', help='Output file for delimited text output', metavar='text-file')
    output_group.add_argument('-e', help='Excel file for output', metavar='Excel-file')
    parser.add_argument('--header', help='Output header line', action='store_true')
    parser.add_argument('--versions', help='Output information about all versions', action='store_true')
    args = parser.parse_args()

    bucket_output = None
    if args.o:
        bucket_output = BucketOutput(OutputType.TextFile, args.header, args.versions, filename=args.o)
    elif args.e:
        bucket_output = BucketOutput(OutputType.Excel, args.header, args.versions, filename=args.e)
    else:
        bucket_output = BucketOutput(OutputType.StandardOutput, args.header, args.versions)

    # s3 = boto3.resource('s3')
    # selected_bucket = s3.Bucket(args.bucket)
    # output_file_objects(selected_bucket, bucket_output)

    repository = Repository(args.bucket)
    print(repository)
